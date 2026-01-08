# app/vacation_form/utils.py
from __future__ import annotations

import os
import calendar
from datetime import date, timedelta
from io import BytesIO
from copy import copy

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage

from app.models import Vacation


def _safe_sheet_title(name: str) -> str:
    bad = [":", "\\", "/", "?", "*", "[", "]"]
    for b in bad:
        name = name.replace(b, "_")
    name = (name or "").strip() or "직원"
    return name[:31]


def _get_fullname(u) -> str:
    ln = (getattr(u, "last_name", "") or "").strip()
    fn = (getattr(u, "first_name", "") or "").strip()
    return (ln + fn).strip() or (getattr(u, "name", "") or "").strip() or (getattr(u, "username", "") or "").strip()


def _normalize_type(t: str) -> str:
    t = (t or "").strip()
    if t in ["반차(전)", "반차(후)"]:
        return "반차"
    return t


def _month_range(year: int, month: int):
    last_day = calendar.monthrange(year, month)[1]
    return date(year, month, 1), date(year, month, last_day)


def _expand_days_in_month(start: date, end: date, first: date, last: date):
    s = max(start, first)
    e = min(end, last)
    d = s
    while d <= e:
        yield d
        d += timedelta(days=1)


def _format_total(x: float) -> str:
    # 2.00 -> 2, 2.50 -> 2.5, 2.75 -> 2.75
    s = f"{x:.2f}".rstrip("0").rstrip(".")
    return s if s else "0"


def _build_b17_leave_summary(user, year: int, month: int) -> str | None:
    """
    [C17] 원하는 형식:
    1월 (연차-5, 12일), (반차-21일), (반반차-26일), (예비군-23일)
    총 2.25일

    ✅ 해당 월에 연차/반차/반반차/병가/예비군/토연차 내역이 하나도 없으면 None 반환
    """
    first, last = _month_range(year, month)

    events = (
        Vacation.query
        .filter(Vacation.user_id == user.id)
        .filter(Vacation.approved == True)
        .filter(Vacation.start_date <= last, Vacation.end_date >= first)
        .all()
    )

    # ✅ 합계 반영(병가/예비군은 표시만 하고 0일로 두는 기존 정책 유지)
    weights = {"연차": 1.0, "반차": 0.5, "반반차": 0.25, "토연차": 0.75}
    ignore_types = {"탄력근무", "근무자"}
    target_types = {"연차", "반차", "반반차", "병가", "예비군", "토연차"}

    days_by_type: dict[str, set[int]] = {}
    total = 0.0

    for e in events:
        t = _normalize_type(getattr(e, "type", ""))
        if (not t) or (t in ignore_types) or (t not in target_types):
            continue

        s = getattr(e, "start_date", None)
        ed = getattr(e, "end_date", None) or s
        if not s:
            continue

        for d in _expand_days_in_month(s, ed, first, last):
            days_by_type.setdefault(t, set()).add(d.day)
            total += weights.get(t, 0.0)

    # ✅ 내역 없으면 시트 생성 X
    if not days_by_type:
        return None

    def _days_str(days: set[int]) -> str:
        d = sorted(days)
        return f"{', '.join(map(str, d))}일"

    order = ["연차", "반차", "반반차", "토연차", "예비군", "병가"]
    parts = []
    for t in order:
        if t in days_by_type:
            parts.append(f"({t}-{_days_str(days_by_type[t])})")

    line1 = f"{month}월 " + ", ".join(parts)
    line2 = f"총 {_format_total(total)}일"
    return f"{line1}\n{line2}"


def build_vacation_forms_xlsx(
    template_path: str,
    users: list,
    *,
    year: int,
    month: int,
    signatures_folder: str | None = None,
    header_signatures: dict[str, str] | None = None,  # ✅ 추가
) -> BytesIO:
    """
    template_path: forms/vacation_form.xlsx
    users: 부서 대상자(User 모델) 리스트
    year/month: B17 휴가기간 계산용
    signatures_folder: (선택) 서명 삽입용 폴더
    """
    wb = load_workbook(template_path)
    ws_tpl = wb.worksheets[0]

    if not users:
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        return bio

    def _sig_path_from_user(u):
        sig = (getattr(u, "signature_image", "") or "").strip()
        if not sig or not signatures_folder:
            return None
        if os.path.isabs(sig) and os.path.exists(sig):
            return sig
        return os.path.join(signatures_folder, os.path.basename(sig))

    def _add_sig(ws, cell_addr: str, sig_path: str, *, width: int, height: int):
        # ✅ 근무표 방식과 동일: 셀 크기 건드리지 않고 add_image만
        if not sig_path or (not os.path.exists(sig_path)):
            return False
        try:
            img = XLImage(sig_path)
            img.width = width
            img.height = height
            ws.add_image(img, cell_addr)
            return True
        except Exception:
            return False

    def fill(ws, u, c17_text: str):
        full_name = _get_fullname(u)
        dept = (getattr(u, "department", "") or "").strip()

        is_super = bool(getattr(u, "is_superadmin", False))
        is_mgr = bool(getattr(u, "is_admin", False)) and (not is_super)
        position = "부서장" if is_mgr else "사원"

        birthday = (getattr(u, "birthday", "") or "").strip()
        phone = (getattr(u, "phone", "") or "").strip()
        address = (getattr(u, "address", "") or "").strip()

        # ✅ 인적사항
        ws["D12"].value = full_name
        ws["F12"].value = dept
        ws["M12"].value = position
        ws["F13"].value = birthday
        ws["M13"].value = phone
        ws["D14"].value = address

        # ✅ [B17] 휴가 기간/내역 요약
        ws["C17"].value = c17_text
        # ✅ [C17] 가운데 정렬 + 줄바꿈 허용
        al = copy(ws["C17"].alignment)
        al.horizontal = "center"
        al.vertical = "center"
        al.wrap_text = True
        ws["C17"].alignment = al

        # ✅ [B25] 신청인         (fullname)(서명)  (공백 9칸, 오른쪽 정렬)
        ws["B25"].value = f"신청인{' ' * 9}{full_name}(서명)"
        al = copy(ws["B25"].alignment)
        al.horizontal = "right"
        ws["B25"].alignment = al

        # ✅ [H25] 직원 개인 서명 이미지
        sig_path = _sig_path_from_user(u)
        if sig_path:
            _add_sig(ws, "H25", sig_path, width=85, height=60)  # 너가 쓰던 값 유지
        
        # ✅ [I3/L3/O3] 결재라인 서명(부서장/부장/병원장) - 시트 공통
        if header_signatures:
            # 부서장(I3)
            p = header_signatures.get("I3")
            if p:
                _add_sig(ws, "I3", p, width=85, height=60)

            # 부장(L3)  (간호부장/행정부장 중 토글된 1명)
            p = header_signatures.get("L3")
            if p:
                _add_sig(ws, "L3", p, width=85, height=60)

            # 병원장(O3)
            p = header_signatures.get("O3")
            if p:
                _add_sig(ws, "O3", p, width=85, height=60)

    eligible: list[tuple[object, str]] = []
    for u in (users or []):
        c17 = _build_b17_leave_summary(u, year, month)
        if c17:  # None이면 스킵
            eligible.append((u, c17))

    if not eligible:
        # 전부 휴가내역 없으면: 파일을 아예 만들지 않는 게 요구사항
        raise ValueError(f"{year}년 {month}월 휴가 내역이 없어 휴가계 시트를 생성하지 않습니다.")

    # ✅ 첫번째 시트: eligible[0]로 채우기
    first, first_c17 = eligible[0]
    ws_tpl.title = _safe_sheet_title(_get_fullname(first))
    fill(ws_tpl, first, first_c17)

    # ✅ 나머지 시트 복제(eligible 기준)
    used = set(wb.sheetnames)
    for u, c17_text in eligible[1:]:
        ws = wb.copy_worksheet(ws_tpl)

        base = _safe_sheet_title(_get_fullname(u))
        title = base
        n = 2
        while title in used:
            title = _safe_sheet_title(f"{base}_{n}")
            n += 1
        ws.title = title
        used.add(title)

        fill(ws, u, c17_text)


    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio
