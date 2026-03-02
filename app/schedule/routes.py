from flask import request, jsonify, send_file, current_app
from flask_login import login_required
from datetime import datetime, date
from app.schedule import schedule_bp
from app.models import User, Vacation, MonthLock, MonthSignToggle, ApprovalRoleUser
from app.schedule.utils import (
    thin_border,
    thin_side,
    uniform_mixed_border,
    find_name_index
)
import calendar
import io
import os
from app import db
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from copy import copy
from openpyxl.drawing.image import Image as XLImage

# --- 세로 굵은선 설정용 ---
THIN = Side(style="thin", color="000000")
MEDIUM = Side(style="medium", color="000000")

def apply_vertical_border(cell, left=False, right=False):
    cell.border = Border(
        left=MEDIUM if left else cell.border.left,
        right=MEDIUM if right else cell.border.right,
        top=cell.border.top,
        bottom=cell.border.bottom
    )

LEFT_MEDIUM_COLS = [1]  # A열 왼쪽 굵은선
RIGHT_MEDIUM_COLS = [2, 33, 34, 35, 36]  # B, AG, AH, AI, AJ 열 오른쪽 굵은선
ADMIN_HEAD_DEPTS = {"도수", "물리치료", "심사과", "원무과", "총무과", "홍보", "진단검사", "영양", "약제부"}
NURSE_HEAD_DEPTS = {"병동", "상담실", "수술실", "외래"}

# =========================================================
# 근무표 자동 생성 (블루프린트 버전)
# URL: /schedule/export/<dept>?year=2025&month=11
# =========================================================
@schedule_bp.route("/export/<dept>")
@login_required
def export_schedule(dept):

    # ====== 기본 날짜 ======
    year = request.args.get("year", type=int, default=datetime.now().year)
    month = request.args.get("month", type=int, default=datetime.now().month)
    last_day = calendar.monthrange(year, month)[1]

    # ====== 폼 파일 로드 ======
    FORM_DIR = current_app.config["FORMS_FOLDER"]
    TEMPLATE_FILE = "gaja_schedule.xlsx"
    template_path = os.path.join(FORM_DIR, TEMPLATE_FILE)

    if not os.path.exists(template_path):
        return jsonify({"error": f"기준 폼이 없습니다: {template_path}"}), 404

    wb = load_workbook(template_path)
    ws = wb[wb.sheetnames[0]]
    
    # ✅ 여기 추가: 시트 이름도 월에 맞게 변경
    ws.title = f"{month}월"

    # 제목 자동 갱신
    ws["A1"] = f"{year}년 {month}월 근무표 (부서: {dept})"

    # ====== 날짜 라벨(C7~) ======
    start_col = 3  # C열부터 날짜
    for day in range(1, last_day + 1):
        col = start_col + (day - 1)
        ws.cell(row=7, column=col).value = day
        weekday = datetime(year, month, day).weekday()
        if weekday == 6:  # 일요일
            ws.cell(row=7, column=col).fill = PatternFill(
                start_color="FFB0B0", end_color="FFB0B0", fill_type="solid"
            )
    # ====== 직원 목록 ======
    employees = (
        User.query.filter_by(department=dept)
        .order_by(User.join_date.asc())
        .all()
    )
    names = [e.name.strip() for e in employees]
    
    # ✅ 추가: user_id -> 근무표 행 index
    id_to_idx = {u.id: i for i, u in enumerate(employees)}

    # ====== 행 복제 ======
    template_row = 8
    if len(names) > 1:
        ws.insert_rows(template_row + 1, len(names) - 1)

    # ====== 복제 + 스타일 ======
    for i, name in enumerate(names):
        target_row = template_row + i

        for col in range(1, 37):  # A~AJ 범위
            src = ws.cell(row=template_row, column=col)
            tgt = ws.cell(row=target_row, column=col)

            if src.has_style:
                try:
                    tgt.font = copy(src.font)
                    tgt.fill = copy(src.fill)
                    tgt.border = copy(src.border)
                    tgt.alignment = copy(src.alignment)
                except:
                    tgt.border = thin_border()
                    tgt.alignment = Alignment(horizontal="center", vertical="center")

            if i > 0:
                uniform_mixed_border(tgt)
                
        # --------------------------------------------------------
        # ⭐ 직원 행 복제 후 — 세로 굵은선(A,B,AG,AH,AI,AJ 열) 복구
        # --------------------------------------------------------

        # A열 왼쪽 굵은선
        apply_vertical_border(ws.cell(target_row, 1), left=True)

        # B, AG, AH, AI, AJ 오른쪽 굵은선
        for col in RIGHT_MEDIUM_COLS:
            apply_vertical_border(ws.cell(target_row, col), right=True)


        # --------------------------
        # 순번(A), 이름(B) 값 설정
        # --------------------------
        ws[f"A{target_row}"].value = i + 1
        ws[f"B{target_row}"].value = name
        ws[f"A{target_row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"B{target_row}"].alignment = Alignment(horizontal="center", vertical="center")

    # ====== 모든 셀 기본 값 채우기 ======
    thin = Side(style="thin", color="000000")
    fill_white = PatternFill("solid", "FFFFFF")
    fill_sunday = PatternFill("solid", "FFB0B0")

    for i, name in enumerate(names):
        row = 8 + i
        for day in range(1, last_day + 1):
            col = 3 + (day - 1)
            cell = ws.cell(row=row, column=col)

            weekday = datetime(year, month, day).weekday()
            if weekday <= 4:  # 평일
                cell.value = "·"
                cell.fill = fill_white
            elif weekday == 5:  # 토요일
                cell.value = "/"
                cell.fill = fill_white
            else:  # 일요일
                cell.value = ""
                cell.fill = fill_sunday

            cell.alignment = Alignment(horizontal="center", vertical="center")

    # ====== 승인된 일정 불러오기 ======
    first_date = date(year, month, 1)
    last_date = date(year, month, last_day)

    events = (
        Vacation.query.filter_by(department=dept)
        .filter(Vacation.approved == True)
        .filter(Vacation.type != "탄력근무")
        .filter(Vacation.start_date >= first_date, Vacation.start_date <= last_date)
        .all()
    )

    # ====== 이벤트 덮어쓰기 ======
    for e in events:
        # ✅ 월이 겹치는지(범위 포함) 체크
        if e.start_date.year != year or e.start_date.month != month:
            continue
            
        # ✅ 근무자: 레거시 데이터(user_id가 등록자일 수 있음) 때문에 name을 먼저 신뢰
        if e.type == "근무자":
            idx = find_name_index((e.name or "").strip(), names)

            # ✅ name으로 못 찾으면 "target_user_id가 있는 경우만" id로 fallback
            if idx is None:
                tuid = getattr(e, "target_user_id", None)
                if tuid:
                    idx = id_to_idx.get(tuid)
                else:
                    # ✅ 퇴사자/매칭불가/레거시(등록자 user_id만 있는 것)는 스킵
                    continue

        else:
            # ✅ 휴가/연차류는 "일정 주인"이 target_user_id 이므로 이걸 우선 사용
            uid = getattr(e, "target_user_id", None) or getattr(e, "user_id", None)
            idx = id_to_idx.get(uid)

            # 레거시 보완: id 매칭이 안되면 name fallback
            if idx is None:
                idx = find_name_index((e.name or "").strip(), names)

        if idx is None:
            continue

        row = 8 + idx
        col = 3 + e.start_date.day - 1

        value = e.type
        if value in ["반차(전)", "반차(후)"]:
            value = "반차"

        weekday = e.start_date.weekday()

        if weekday == 5:  # 토요일
            if value == "근무자":
                value = "·"
            elif value == "토연차":
                value = "토연차"   # ← 토연차 그대로 표시
            else:
                value = "/"        # ← 나머지 토요일 일정만 "/"


        cell = ws.cell(row=row, column=col)
        cell.value = value

        # 긴 텍스트 자동 축소
        if len(str(value)) >= 3:
            cell.alignment = Alignment(
                shrinkToFit=True, horizontal="center", vertical="center"
            )
        else:
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 테두리 보정
        medium = Side(style="medium", color="000000")
        cell.border = Border(left=thin, right=thin, top=medium, bottom=medium)

    # ====== 합계 (AI, AJ) ======
    weights = {"연차": 1.0, "반차": 0.5, "반반차": 0.25, "토연차": 0.75}
    sick_types = ["병가", "예비군"]

    for i, user in enumerate(employees):
        row = 8 + i

        # ✅ 이 직원의 이벤트만 선택: 휴가/연차는 target_user_id(주인) 우선
        user_events = [
            v for v in events
            if (
                getattr(v, "target_user_id", None) == user.id
                or (getattr(v, "target_user_id", None) in (None, 0) and v.user_id == user.id)
            )
            and v.start_date.month == month
        ]

        # 연차 합계 (반차 합치기 / 토연차 0.75 반영)
        total_leave = sum(
            weights.get(
                "반차" if v.type in ["반차(전)", "반차(후)"] else v.type,
                0
            )
            for v in user_events
        )

        # 병가 / 예비군
        total_sick = sum(1 for v in user_events if v.type in sick_types)

        # AI (연차)
        ai = ws[f"AI{row}"]
        ai.value = total_leave
        ai.alignment = Alignment(horizontal="center", vertical="center", shrinkToFit=True)

        # AJ (병가/예비군)
        aj = ws[f"AJ{row}"]
        aj.value = total_sick
        aj.alignment = Alignment(horizontal="center", vertical="center")



    # ====== 인쇄 설정 ======
    last_row = 8 + len(names) - 1
    ws.print_area = f"A1:AJ{last_row}"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.2
    ws.page_margins.right = 0.2
    ws.page_margins.top = 0.3
    ws.page_margins.bottom = 0.3
    ws.page_setup.horizontalCentered = True
    ws.page_setup.verticalCentered = True
    ws.print_title_rows = "1:7"

    # =========================================================
    # ✅ (확정된 달이면) 서명 삽입
    #  - AA3: 확정한 중간관리자 서명(기존 유지)
    #  - AD3: 부장(행정부장/간호부장) 서명 (토글 ON + 부서 매핑)
    #  - AG3: 병원장 서명 (토글 ON)
    #  - 단, "부서 확정(locked)" 된 부서만 적용
    # =========================================================
    dept_key = (dept or "").strip()
    lk = MonthLock.query.filter_by(department=dept_key, year=year, month=month).first()

    def _sig_path_from_user(u: User):
        sig = (getattr(u, "signature_image", "") or "").strip() if u else ""
        if not sig:
            return None

        # 1) DB에 절대경로가 저장된 경우
        if os.path.isabs(sig) and os.path.exists(sig):
            return sig

        # 2) DB에 "폴더/파일명" 형태가 저장된 경우 → 파일명만 뽑아서 SIGNATURES_FOLDER에 붙임
        sig_base = os.path.basename(sig)
        return os.path.join(current_app.config["SIGNATURES_FOLDER"], sig_base)
    

    def _add_sig(ws, cell_addr: str, sig_path: str):
        if not sig_path or (not os.path.exists(sig_path)):
            return False
        try:
            img = XLImage(sig_path)
            # ✅ 기존 부서장 서명 크기와 동일하게
            img.width = 100
            img.height = 64
            ws.add_image(img, cell_addr)
            return True
        except Exception as e:
            current_app.logger.exception(
                "SIGNATURE INSERT FAILED: dept=%s y=%s m=%s cell=%s path=%s err=%s",
                dept_key, year, month, cell_addr, sig_path, repr(e)
            )
            return False
        
    def _get_role_user(role: str):
        rec = ApprovalRoleUser.query.filter_by(role=role).first()
        return rec.user if rec else None


    # ✅ 확정된 달(locked)인 경우에만 서명 로직 실행
    if lk and lk.locked:

        # -----------------------------------------------------
        # 1) (기존) 확정한 중간관리자 서명 → AA3
        # -----------------------------------------------------
        if lk.locked_by:
            signer = db.session.get(User, int(lk.locked_by))
            sig_path = _sig_path_from_user(signer)
            if sig_path and os.path.exists(sig_path):
                _add_sig(ws, "AA3", sig_path)
            else:
                current_app.logger.warning(
                    "Month locked but manager signature missing. dept=%s %04d-%02d locked_by=%s path=%s",
                    dept_key, year, month, lk.locked_by, sig_path
                )

        # -----------------------------------------------------
        # 2) 결재라인 토글 상태 조회 (월 단위)
        # -----------------------------------------------------
        tg = MonthSignToggle.query.filter_by(year=year, month=month).first()
        director_on = bool(tg and tg.director_on)
        admin_head_on = bool(tg and tg.admin_head_on)
        nurse_head_on = bool(tg and tg.nurse_head_on)

        # -----------------------------------------------------
        # 3) 병원장 서명 → AG3 (토글 ON일 때만)
        # -----------------------------------------------------
        if director_on:
            director = _get_role_user("director")
            director_path = _sig_path_from_user(director)
            if not _add_sig(ws, "AG3", director_path):
                current_app.logger.warning(
                    "Director toggle ON but signature missing. dept=%s %04d-%02d path=%s",
                    dept_key, year, month, director_path
                )

        # -----------------------------------------------------
        # 4) 부장 서명 → AD3 (부서에 따라 행정부장/간호부장, 그리고 토글 ON 조건)
        # -----------------------------------------------------
        head_name = None
        if dept_key in ADMIN_HEAD_DEPTS:
            head_name = "행정부장" if admin_head_on else None
        elif dept_key in NURSE_HEAD_DEPTS:
            head_name = "간호부장" if nurse_head_on else None

        if head_name:
            head_user = _get_role_user("admin_head") if head_name == "행정부장" else _get_role_user("nurse_head")
            head_path = _sig_path_from_user(head_user)
            if not _add_sig(ws, "AD3", head_path):
                current_app.logger.warning(
                    "Head toggle ON but signature missing. head=%s dept=%s %04d-%02d path=%s",
                    head_name, dept_key, year, month, head_path
                )


    # ====== 파일 저장 후 전송 ======

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{dept}_근무표_{year}_{month:02d}_{stamp}.xlsx"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    resp = send_file(output, as_attachment=True, download_name=filename)

    # ✅ 캐시 방지(브라우저가 이전 다운로드를 재사용하는 문제 방지)
    resp.headers["Cache-Control"] = "no-store, max-age=0"
    resp.headers["Pragma"] = "no-cache"
    resp.headers["Expires"] = "0"
    return resp

